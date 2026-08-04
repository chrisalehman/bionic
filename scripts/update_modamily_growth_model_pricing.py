#!/usr/bin/env python3
"""Create the Aug. 16 pricing-launch growth model from the immutable source."""

from __future__ import annotations

import argparse
import os
import tempfile
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


DEFAULT_SOURCE = Path(
    "/Users/ivanfatovic/Claude-Work/Outputs/modamily-growth-model.xlsx"
)
DEFAULT_OUTPUT = Path(
    "/Users/ivanfatovic/Claude-Work/Outputs/"
    "modamily-growth-model-pricing-launch-2026-08-16.xlsx"
)

BLUE = "FF0000FF"
BLACK = "FF000000"

SCORECARD_HEADERS = [
    "Week ending",
    "Platform",
    "Pricing cohort",
    "Gross cash collected ($)",
    "Refunds ($)",
    "Fees ($)",
    "Net cash ($)",
    "Normalized new MRR ($)",
    "Active payers",
    "Eligible paywall viewers",
    "Paywall views",
    "Paid starts",
    "Weekly starts",
    "Monthly starts",
    "Quarterly starts",
    "Annual starts",
    "U.S. paid starts",
    "Non-U.S. paid starts",
    "Country unknown",
    "Net revenue / eligible viewer ($)",
    "28d rolling net cash ($)",
    "On track vs cash target?",
]

PAYER_MIX_SOURCE = (
    "Source: User's June 2-August 2 Stripe transaction reconstruction, converted "
    "from charge frequency into an estimated active-payer mix. The 38% annual "
    "share is extrapolated from only four annual charges and is a wide estimate. "
    "Quarterly and monthly shares are more reliable."
)


def copy_style(source, destination) -> None:
    destination._style = copy(source._style)
    destination.number_format = source.number_format
    destination.alignment = copy(source.alignment)
    destination.protection = copy(source.protection)


def set_font_color(cell, color: str) -> None:
    font = copy(cell.font)
    font.color = color
    font.name = "Arial"
    cell.font = font


def update_pricing_ladder(workbook) -> None:
    ws = workbook["Pricing Ladder"]

    # Donors remain invariant after transformation so T(T(x)) has stable styles.
    mix_label_style = copy(ws["A2"]._style)
    plan_label_style = copy(ws["A8"]._style)
    currency_formula_style = copy(ws["D8"]._style)
    currency_input_style = copy(ws["C8"]._style)
    percent_input_style = copy(ws["G8"]._style)
    formula_percent_style = copy(ws["F8"]._style)

    ws["H4"] = "Normalized new-plan mix"
    copy_style(ws["G4"], ws["H4"])
    ws.column_dimensions["H"].width = 23

    payer_shares = [0.38, 0.27, 0.23, 0.08]
    for row, share in zip(range(5, 9), payer_shares):
        ws.cell(row=row, column=7).value = share
        ws.cell(row=row, column=7)._style = copy(percent_input_style)
        set_font_color(ws.cell(row=row, column=7), BLUE)
        ws.cell(row=row, column=7).comment = Comment(PAYER_MIX_SOURCE, "Codex")

        normalized = ws.cell(row=row, column=8)
        normalized.value = f"=G{row}/SUM($G$5:$G$8)"
        normalized._style = copy(formula_percent_style)
        normalized.number_format = "0.0%"
        set_font_color(normalized, BLACK)

    ws["A9"] = "Legacy"
    ws["A9"]._style = plan_label_style
    ws["B9"] = 32.0
    ws["B9"]._style = currency_input_style
    set_font_color(ws["B9"], BLUE)
    ws["B9"].comment = Comment(
        "Assumption from the user: legacy $29.99/$32.99 plans are modeled at "
        "approximately $32 per payer per month. Edit this yellow/blue input; "
        "the monthly-equivalent cell in D9 is derived from it.",
        "Codex",
    )
    ws["C9"] = None
    ws["D9"] = "=B9"
    ws["D9"]._style = currency_formula_style
    set_font_color(ws["D9"], BLACK)
    ws["D9"].comment = None
    ws["E9"] = None
    ws["F9"] = None
    ws["G9"] = 0.04
    ws["G9"]._style = percent_input_style
    set_font_color(ws["G9"], BLUE)
    ws["G9"].comment = Comment(PAYER_MIX_SOURCE, "Codex")
    ws["H9"] = None

    ws["A10"] = "Payer mix check (must = 100%)"
    ws["A10"]._style = mix_label_style
    ws["G10"] = "=SUM(G5:G9)"
    ws["G10"]._style = copy(formula_percent_style)
    set_font_color(ws["G10"], BLACK)
    ws["H10"] = "=SUM(H5:H8)"
    ws["H10"]._style = copy(formula_percent_style)
    set_font_color(ws["H10"], BLACK)
    for ref in ("G10", "H10"):
        font = copy(ws[ref].font)
        font.bold = True
        ws[ref].font = font

    ws["D12"] = "=SUMPRODUCT(D5:D9,G5:G9)"
    set_font_color(ws["D12"], BLACK)
    ws["E13"] = "=SUMPRODUCT(E5:E8,H5:H8)"
    set_font_color(ws["E13"], BLACK)
    ws["A13"] = "New ladder (normalized new-plan mix)"
    ws["A14"] = "New-plan normalized ARPU vs current-base ARPU"
    ws["E14"] = "=E13/D12-1"
    set_font_color(ws["E14"], BLACK)

    ws["A17"] = (
        "- Payer mix is estimated from the user's June 2-August 2 Stripe transaction "
        "reconstruction after converting charge frequency into active payers. Annual "
        "38% is a wide estimate based on four charges; validate against subscription counts."
    )
    ws["A18"] = (
        "- Longer terms raise prepaid cash and expected lifetime but lower monthly-equivalent "
        "revenue versus monthly. The June cash step was not purely an ARPU change."
    )
    ws["A19"] = (
        "- Quarterly and annual purchases collect cash upfront; amortize each term in normalized "
        "MRR so cash timing is not mistaken for recurring growth."
    )
    ws["A20"] = (
        "- Judge pricing cells on net revenue per eligible paywall viewer; also track term mix, "
        "refunds, complaints, normalized MRR, and collected cash separately."
    )
    ws["A21"] = (
        "- New users only; grandfather existing subscribers. App Store and Play price changes "
        "must be coordinated with the Stripe launch cohort."
    )
    ws["A22"] = (
        "- The $49.99 monthly anchor is at the cited competitor level; annual $199.99 is "
        "$16.67/month equivalent and should be described as prepaid value, not monthly cash."
    )


def update_path_to_30k(workbook) -> None:
    ws = workbook["Path to $30K"]
    ws["A1"] = "Path to $30K normalized MRR - the four-lever unit model"
    ws["A2"] = (
        "Steady-state normalized MRR = signups/mo x conversion % x avg paying lifetime "
        "(months) x normalized $/payer/mo. Cash collected is tracked separately in the "
        "Weekly Scorecard."
    )
    ws["A8"] = "Normalized revenue per payer per month"
    ws["D8"] = "='Pricing Ladder'!$E$13"
    ws["E8"] = (
        "Linked to Pricing Ladder normalized new-subscriber MRR per payer; excludes legacy "
        "and amortizes prepaid terms over their billing periods."
    )
    ws["A13"] = "STEADY-STATE NORMALIZED MRR"
    ws["A14"] = "Hits $30K normalized MRR target?"


def update_weekly_scorecard(workbook) -> None:
    ws = workbook["Weekly Scorecard"]

    # These donors keep the same semantic role after each transformation.
    header_style = copy(ws["A4"]._style)
    # B5 remains a general-format manual text input; A5 becomes a date input.
    manual_text_style = copy(ws["B5"]._style)
    formula_currency_style = copy(workbook["Pricing Ladder"]["D5"]._style)
    note_style = copy(ws["A6"]._style)

    for row in range(4, 7):
        for column in range(1, 23):
            ws.cell(row=row, column=column).value = None
            ws.cell(row=row, column=column).comment = None

    ws["A1"] = "Weekly Cash Scorecard - update every Monday (from the playbook)"
    ws["A2"] = (
        "Fill blue-font cells weekly; one row per platform and pricing cohort. The target "
        "view is 28-day rolling net cash; normalized new MRR is reported separately."
    )

    for column, header in enumerate(SCORECARD_HEADERS, start=1):
        cell = ws.cell(row=4, column=column, value=header)
        cell._style = copy(header_style)

    # Preserve the original example's known cash fields without manufacturing term mix.
    manual_values = {
        "A5": date(2026, 8, 8),
        "B5": "Stripe",
        "C5": "Pre-launch / mixed",
        "D5": 829,
        "E5": 0,
        "F5": 34,
        "I5": None,
        "J5": 57000,
        "K5": None,
        "L5": 25,
        "M5": None,
        "N5": None,
        "O5": None,
        "P5": None,
        "Q5": None,
        "R5": None,
        "S5": None,
    }
    currency_columns = {4, 5, 6}
    for ref, value in manual_values.items():
        cell = ws[ref]
        cell.value = value
        cell._style = copy(manual_text_style)
        if cell.column in currency_columns:
            cell.number_format = '\\$#,##0.00;("$"#,##0.00);-'
        set_font_color(cell, BLUE)
    ws["A5"].number_format = "yyyy-mm-dd"

    formula_values = {
        "G5": "=D5-E5-F5",
        "H5": (
            "=M5*'Pricing Ladder'!$E$8+N5*'Pricing Ladder'!$E$7+"
            "O5*'Pricing Ladder'!$E$6+P5*'Pricing Ladder'!$E$5"
        ),
        "T5": "=IFERROR(G5/J5,0)",
        "U5": '=SUMIFS($G:$G,$B:$B,B5,$A:$A,">="&A5-27,$A:$A,"<="&A5)',
        "V5": '=IF(U5>=Assumptions!$B$20,"YES","NO")',
    }
    for ref, formula in formula_values.items():
        cell = ws[ref]
        cell.value = formula
        cell._style = copy(formula_currency_style)
        set_font_color(cell, BLACK)

    for ref in ["G5", "H5", "T5", "U5"]:
        ws[ref].number_format = '\\$#,##0.00;("$"#,##0.00);-'
    ws["V5"].number_format = "General"

    ws["D5"].comment = Comment(
        "Gross cash collected in the week before refunds and processor/store fees.",
        "Codex",
    )
    ws["H5"].comment = Comment(
        "Formula amortizes new weekly, monthly, quarterly, and annual starts into a "
        "monthly-equivalent run rate using Pricing Ladder monthly equivalents.",
        "Codex",
    )
    ws["J5"].comment = Comment(
        "Eligible paywall viewers are the denominator for the pricing decision metric; "
        "do not substitute total contacts or charge count.",
        "Codex",
    )

    ws["A6"] = (
        "Template: replace the blue example inputs with weekly platform/cohort data; leave "
        "black formula cells intact. Term-start and geography counts must reconcile to paid starts."
    )
    ws["A6"]._style = note_style

    widths = [
        13,
        12,
        20,
        24,
        14,
        12,
        15,
        24,
        15,
        25,
        14,
        13,
        15,
        15,
        17,
        14,
        16,
        20,
        17,
        31,
        25,
        25,
    ]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.row_dimensions[4].height = 42
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:V5"


def transform(source: Path, output: Path) -> None:
    source = source.expanduser().resolve()
    output = output.expanduser().resolve()
    if source == output:
        raise ValueError("Source workbook is immutable; output must use a different path")
    if not source.exists():
        raise FileNotFoundError(source)

    workbook = load_workbook(source, data_only=False)
    try:
        update_pricing_ladder(workbook)
        update_path_to_30k(workbook)
        update_weekly_scorecard(workbook)
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

        output.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            prefix=f".{output.stem}-", suffix=".xlsx", dir=output.parent, delete=False
        ) as handle:
            temporary_path = Path(handle.name)
        try:
            workbook.save(temporary_path)
            os.replace(temporary_path, output)
        finally:
            temporary_path.unlink(missing_ok=True)
    finally:
        workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    transform(arguments.source, arguments.output)
    print(arguments.output)
