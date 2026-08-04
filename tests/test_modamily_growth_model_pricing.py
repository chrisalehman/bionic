import os
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.update_modamily_growth_model_pricing import transform


SOURCE_PATH = Path(
    "/Users/ivanfatovic/Claude-Work/Outputs/modamily-growth-model.xlsx"
)
OUTPUT_PATH = Path(
    "/Users/ivanfatovic/Claude-Work/Outputs/"
    "modamily-growth-model-pricing-launch-2026-08-16.xlsx"
)

EXPECTED_SCORECARD_HEADERS = [
    "Week ending",
    "Platform",
    "Pricing cohort",
    "Gross cash collected ($)",
    "Refunds ($)",
    "Fees ($)",
    "Net cash ($)",
    "Normalized new MRR ($)",
    "Active payers",
    "Eligible paywall viewers",
    "Paywall views",
    "Paid starts",
    "Weekly starts",
    "Monthly starts",
    "Quarterly starts",
    "Annual starts",
    "U.S. paid starts",
    "Non-U.S. paid starts",
    "Country unknown",
    "Net revenue / eligible viewer ($)",
    "28d rolling net cash ($)",
    "On track vs cash target?",
]


def normalized_formula(cell_ref):
    return str(cell_ref).replace(" ", "").upper()


class PricingLaunchWorkbookTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        selected = os.environ.get("MODAMILY_MODEL_PATH", str(OUTPUT_PATH))
        cls.path = Path(selected)
        if not cls.path.exists():
            raise AssertionError(f"Workbook does not exist: {cls.path}")
        # Formula strings and cached values cannot be obtained from one load.
        cls.formulas = load_workbook(cls.path, data_only=False)
        cls.values = load_workbook(cls.path, data_only=True)

    @classmethod
    def tearDownClass(cls):
        cls.formulas.close()
        cls.values.close()

    def test_pricing_ladder_uses_corrected_payer_mix_and_legacy_row(self):
        ws = self.formulas["Pricing Ladder"]
        self.assertEqual(
            [ws.cell(row=row, column=1).value for row in range(5, 10)],
            [
                "Annual (12m)",
                "Quarterly (3m)",
                "Monthly (1m)",
                "Weekly (1w)",
                "Legacy",
            ],
        )
        self.assertEqual(
            [ws.cell(row=row, column=7).value for row in range(5, 10)],
            [0.38, 0.27, 0.23, 0.08, 0.04],
        )
        self.assertEqual(ws["B9"].value, 32.0)
        self.assertEqual(normalized_formula(ws["D9"].value), "=B9")

    def test_legacy_monthly_assumption_has_one_documented_input(self):
        formula_ws = self.formulas["Pricing Ladder"]
        value_ws = self.values["Pricing Ladder"]
        input_cell = formula_ws["B9"]
        derived_cell = formula_ws["D9"]

        self.assertEqual(input_cell.value, 32.0)
        self.assertEqual(input_cell.font.color.type, "rgb")
        self.assertEqual(input_cell.font.color.rgb, "FF0000FF")
        self.assertEqual(input_cell.fill.fgColor.rgb, "FFFFFF00")
        self.assertIsNotNone(input_cell.comment)
        self.assertIn("legacy", input_cell.comment.text.lower())
        self.assertIn("approximately $32", input_cell.comment.text.lower())

        self.assertEqual(normalized_formula(derived_cell.value), "=B9")
        self.assertEqual(derived_cell.font.color.type, "rgb")
        self.assertEqual(derived_cell.font.color.rgb, "FF000000")
        self.assertNotEqual(derived_cell.fill.fgColor.rgb, "FFFFFF00")
        self.assertEqual(value_ws["D9"].value, 32.0)

    def test_new_plan_mix_normalizes_to_one_and_excludes_legacy(self):
        formula_ws = self.formulas["Pricing Ladder"]
        value_ws = self.values["Pricing Ladder"]
        for row in range(5, 9):
            self.assertTrue(
                str(formula_ws.cell(row=row, column=8).value).startswith("="),
                f"H{row} must be a formula",
            )
        self.assertIsNone(formula_ws["H9"].value)
        self.assertAlmostEqual(
            sum(value_ws.cell(row=row, column=8).value for row in range(5, 9)),
            1.0,
            places=10,
        )

    def test_new_ladder_blended_arpu_excludes_legacy(self):
        ws = self.formulas["Pricing Ladder"]
        self.assertEqual(
            normalized_formula(ws["E13"].value),
            "=SUMPRODUCT(E5:E8,H5:H8)",
        )
        self.assertAlmostEqual(self.values["Pricing Ladder"]["E13"].value, 35.1654, places=4)

    def test_path_to_30k_is_explicitly_normalized_mrr(self):
        ws = self.formulas["Path to $30K"]
        self.assertEqual(
            normalized_formula(ws["D8"].value),
            "='PRICINGLADDER'!$E$13",
        )
        self.assertIn("normalized MRR", str(ws["A1"].value))
        self.assertIn("normalized", str(ws["E8"].value).lower())

    def test_weekly_scorecard_has_exact_launch_headers(self):
        ws = self.formulas["Weekly Scorecard"]
        self.assertEqual(
            [ws.cell(row=4, column=column).value for column in range(1, 23)],
            EXPECTED_SCORECARD_HEADERS,
        )

    def test_weekly_scorecard_uses_cash_formulas(self):
        ws = self.formulas["Weekly Scorecard"]
        self.assertEqual(normalized_formula(ws["G5"].value), "=D5-E5-F5")
        self.assertEqual(normalized_formula(ws["T5"].value), "=IFERROR(G5/J5,0)")
        self.assertEqual(
            normalized_formula(ws["U5"].value),
            '=SUMIFS($G:$G,$B:$B,B5,$A:$A,">="&A5-27,$A:$A,"<="&A5)',
        )
        self.assertEqual(
            normalized_formula(ws["V5"].value),
            '=IF(U5>=ASSUMPTIONS!$B$20,"YES","NO")',
        )
        self.assertTrue(str(ws["T5"].value).startswith("="))
        self.assertIn("cash", str(ws["A1"].value).lower())

    def test_weekly_scorecard_preserves_input_and_formula_row_formats(self):
        ws = self.formulas["Weekly Scorecard"]
        manual_refs = [
            "A5",
            "B5",
            "C5",
            "D5",
            "E5",
            "F5",
            "I5",
            "J5",
            "K5",
            "L5",
            "M5",
            "N5",
            "O5",
            "P5",
            "Q5",
            "R5",
            "S5",
        ]
        for ref in manual_refs:
            self.assertEqual(ws[ref].font.color.type, "rgb", ref)
            self.assertEqual(ws[ref].font.color.rgb, "FF0000FF", ref)
        for ref in ["G5", "H5", "T5", "U5", "V5"]:
            self.assertEqual(ws[ref].font.color.type, "rgb", ref)
            self.assertEqual(ws[ref].font.color.rgb, "FF000000", ref)
            self.assertNotEqual(ws[ref].fill.fgColor.rgb, "FFFFFF00", ref)

        self.assertEqual(ws["A5"].number_format.replace("\\", ""), "yyyy-mm-dd")
        for ref in ["D5", "E5", "F5", "G5", "H5", "T5", "U5"]:
            self.assertIn("$", ws[ref].number_format, ref)
        self.assertEqual(ws["V5"].number_format, "General")

    def test_payer_mix_source_and_annual_uncertainty_are_documented(self):
        ws = self.formulas["Pricing Ladder"]
        comments = " ".join(
            cell.comment.text for row in ws.iter_rows() for cell in row if cell.comment
        ).lower()
        self.assertIn("june 2", comments)
        self.assertIn("august 2", comments)
        self.assertIn("four annual charges", comments)
        self.assertIn("wide estimate", comments)
        cautions = " ".join(
            str(ws.cell(row=row, column=1).value or "") for row in range(16, 24)
        ).lower()
        self.assertNotIn("the price change is the arpu lever", cautions)
        self.assertIn("prepaid cash", cautions)

    def test_every_formula_has_a_clean_cached_value_after_recalculation(self):
        error_tokens = {
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        }
        formula_locations = []
        for sheet_name in self.formulas.sheetnames:
            formula_sheet = self.formulas[sheet_name]
            cached_sheet = self.values[sheet_name]
            for row in formula_sheet.iter_rows():
                for cell in row:
                    if not (
                        cell.data_type == "f"
                        or (isinstance(cell.value, str) and cell.value.startswith("="))
                    ):
                        continue
                    location = f"{sheet_name}!{cell.coordinate}"
                    formula_locations.append(location)
                    cached_value = cached_sheet[cell.coordinate].value
                    self.assertIsNotNone(cached_value, f"{location} lacks a cached value")
                    if isinstance(cached_value, str):
                        matched = [token for token in error_tokens if token in cached_value]
                        self.assertFalse(matched, f"{location} contains {matched}: {cached_value}")

        self.assertIn("Weekly Scorecard!H5", formula_locations)
        self.assertGreater(len(formula_locations), 200)

    def test_transformation_is_semantically_idempotent(self):
        def workbook_snapshot(path):
            workbook = load_workbook(path, data_only=False)
            try:
                snapshot = {}
                for sheet in workbook.worksheets:
                    cells = {}
                    for row in sheet.iter_rows():
                        for cell in row:
                            comment = None
                            if cell.comment:
                                comment = (cell.comment.text, cell.comment.author)
                            cells[cell.coordinate] = (
                                cell.value,
                                tuple(cell._style) if cell.has_style else None,
                                comment,
                            )
                    columns = {
                        key: (dimension.width, dimension.hidden, dimension.outlineLevel)
                        for key, dimension in sheet.column_dimensions.items()
                    }
                    rows = {
                        key: (dimension.height, dimension.hidden, dimension.outlineLevel)
                        for key, dimension in sheet.row_dimensions.items()
                    }
                    snapshot[sheet.title] = {
                        "bounds": (sheet.max_row, sheet.max_column),
                        "cells": cells,
                        "columns": columns,
                        "rows": rows,
                        "freeze": str(sheet.freeze_panes),
                        "filter": sheet.auto_filter.ref,
                    }
                return snapshot
            finally:
                workbook.close()

        with tempfile.TemporaryDirectory(prefix="modamily-model-idempotence-") as directory:
            first = Path(directory) / "first.xlsx"
            second = Path(directory) / "second.xlsx"
            transform(SOURCE_PATH, first)
            transform(first, second)
            self.assertEqual(workbook_snapshot(first), workbook_snapshot(second))

    def test_formulas_outside_named_sheets_are_preserved(self):
        if self.path.resolve() == SOURCE_PATH.resolve():
            self.skipTest("preservation comparison applies to transformed output")
        source = load_workbook(SOURCE_PATH, data_only=False)
        try:
            named = {"Pricing Ladder", "Path to $30K", "Weekly Scorecard"}
            for sheet_name in source.sheetnames:
                if sheet_name in named:
                    continue
                before = {
                    cell.coordinate: cell.value
                    for row in source[sheet_name].iter_rows()
                    for cell in row
                    if cell.data_type == "f"
                }
                after = {
                    cell.coordinate: cell.value
                    for row in self.formulas[sheet_name].iter_rows()
                    for cell in row
                    if cell.data_type == "f"
                }
                self.assertEqual(after, before, f"formulas changed on {sheet_name}")
        finally:
            source.close()


if __name__ == "__main__":
    unittest.main()
